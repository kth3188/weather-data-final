from openpyxl import load_workbook

class WeatherAPI:
    def __init__(self):
        self.coordinates = self._load_coordinates()

    def _load_coordinates(self):
        """엑셀 파일에서 좌표 데이터를 로드합니다."""
        coordinates = {}
        wb = load_workbook('data/coordinates.xlsx')
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            location_key = f"{row[0]} {row[1]}"
            if row[2]:  # 3단계가 있는 경우
                location_key += f" {row[2]}"
            coordinates[location_key] = {'nx': row[3], 'ny': row[4]}
        
        return coordinates

    def get_coordinates(self, location):
        """위치에 해당하는 좌표를 반환합니다."""
        if location not in self.coordinates:
            raise ValueError(f"Invalid location: {location}")
        return self.coordinates[location]

    def get_weather(self, location):
        """특정 위치의 날씨 정보를 조회합니다."""
        coords = self.get_coordinates(location)
        # ... rest of the existing get_weather code ... 